#####################################################
# OBJECTIFS DU SCRIPT :                         #
# Créer un fichier excel                        #
# Le remplir de valeurs                         #
# Ajouter des conditions de mise en forme       #
#####################################################

#####################################################
# 1 - Import des librairies                     #
#####################################################

import os
import webbrowser
from typing import Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # , Color
from qgis.core import QgsProject
from qgis.gui import QgsMessageBar
from qgis.PyQt.QtCore import NULL
from qgis.PyQt.QtWidgets import (
    QDialog,
    QDialogButtonBox,
    QLabel,
    QSizePolicy,
    QVBoxLayout,
)

LR_COLORS = {
    "EX": "000000",
    "EW": "3d1851",
    "RE": "5b1a62",
    "CR": "d20019",
    "EN": "fabf00",
    "VU": "ffed00",
    "NT": "faf2c7",
    "LC": "78b747",
    "DD": "d4d4d4",
}


def wc_hex_is_light(color):
    """Set font color depending on fill darkness"""
    red = int(color[0:2], 16)
    green = int(color[2:4], 16)
    blue = int(color[4:6], 16)
    brightness = ((red * 299) + (green * 587) + (blue * 114)) / 1000
    return "000000" if brightness > 155 else "ffffff"


class SuccessDialog(QDialog):
    def __init__(self):
        QDialog.__init__(self)
        self.bar = QgsMessageBar()
        self.bar.setSizePolicy(QSizePolicy.Maximum, QSizePolicy.Fixed)
        self.setWindowTitle("EXPORT RÉUSSI !")
        self.setLayout(QVBoxLayout())
        self.layout().setContentsMargins(0, 0, 0, 0)
        message = QLabel(
            f" Le fichier Excel a bien été enregistré dans {os.path.expanduser('~')}/QGIS_exports "
        )
        self.buttonbox = QDialogButtonBox(QDialogButtonBox.Ok)
        self.buttonbox.accepted.connect(self.close)
        self.layout().addWidget(message)
        self.layout().addWidget(self.buttonbox)
        self.layout().addWidget(self.bar)


#####################################################
# >> 2 - Gestion des données                       #
#####################################################

# >> Création du fichier excel
wb = Workbook()
# >> Sélection de "feuille" active 'worksheet'
ws = wb.active

# >> Alimentation du fichier excel avec les données de la table attributaire
# Récupération de la couche
layer_id = "[%@layer_id%]"
layer = QgsProject().instance().mapLayer(layer_id)

# Ajout de l'entête
fields = layer.fields()
fields_row = []
for i, field in enumerate(fields):
    fields_row.append(field.name())
# print(fields_row)
ws.append(fields_row)
# Ajout des entités (lignes)
features = layer.getFeatures()
for feature in features:
    feature_row = []
    for attribute in range(len(fields_row)):
        if feature[attribute] == NULL:
            feature_row.append("")
        else:
            feature_row.append(feature[attribute])
    # print(feature_row)
    ws.append(feature_row)
#####################################################
# >> 3 - Mise en forme du fichier                  #
#####################################################

# >> 3.1 - Mise en italique des cases 'Nom scientifique'

# >> Définition du style
italic_grey_font = Font(color="606060", italic=True)
# >> Rechercher la colonne "Nom scientifique"
for col in ws["1:1"]:
    if col.value == "Nom scientifique":
        # Si on trouve une colonne référente alors :
        range_nom_ver = col
        ref_nom_vern = range_nom_ver.column_letter + ":" + range_nom_ver.column_letter
        for cell in ws[ref_nom_vern]:
            cell.font = italic_grey_font
            cell.alignment = Alignment(horizontal="center")

# >> 3.2 - Couleur sur les cases de type statut


# >> Définition de la fonction d'application des couleurs selon le statut
def color_statut_style(x):
    """set RedList color style"""
    for lr_cell in ws[x]:
        if lr_cell.value in LR_COLORS.keys():
            color = LR_COLORS[lr_cell.value]
            lr_cell.fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )
            lr_cell.font = Font(color=wc_hex_is_light(color))


# >> Recherche des colonnes de type statut
# Rechercher la colonne "LR France", "LR Rhône-Alpes", "LR Auvergne" # d'autres colonnes à prévoir
for col in ws["1:1"]:
    if col.value.startswith("LR "):
        # Si on trouve une colonne référente alors :
        range_statut = col
        ref_statut = range_statut.column_letter + ":" + range_statut.column_letter
        color_statut_style(ref_statut)

# >> 3.3 - Style général des colonnes

# >> Mise en gras des noms de colonnes
for cell in ws["1:1"]:
    cell.font = Font(bold=True, italic=False, vertAlign=None, color="ffffff", size=12)
    cell.fill = PatternFill(start_color="0076bd", end_color="0076bd", fill_type="solid")

# >> Mise en forme de la largeur des colonnes
dims: Dict[str, int] = {}
for row in ws.rows:
    for cell in row:
        if cell.value:
            dims[cell.column_letter] = max(
                (dims.get(cell.column_letter, 0), len(str(cell.value)))
            )
for col, value in dims.items():
    if value < 13:
        ws.column_dimensions[col].width = 13
    else:
        ws.column_dimensions[col].width = value


# >> Mise en forme des bordures du tableau
# Définition d'une fonction qui parcourt les cellules et applique le style de bordure choisie
def set_border(ws, cell_range):
    border = Border(bottom=Side(border_style="thin", color="0076bd"))
    for line in ws[cell_range]:
        for rcell in line:
            rcell.border = border


# Obtenir les dimensions du tableau
full_dim = ws.dimensions
# Appliquer la mise en forme des bordures sur l'ensemble du tableau
set_border(ws, "A1:Z3275")

#####################################################
# >> 4 - Enregistrement du résultat final          #
#####################################################

# Sauvegarde du fichier
path = f"{os.path.expanduser('~')}/QGIS_exports"
if not os.path.exists(path):
    os.mkdir(path)

wb.save(f"{path}/{layer.name()}.xlsx")

webbrowser.open(os.path.realpath(path))
success_dialog = SuccessDialog()
success_dialog.show()
